#!/usr/bin/env python3
"""Build Shopify blog redirect batches from a Prompt 6 consolidation sheet.

The expected input columns are:
- keep_or_redirect
- redirect_from_slug
- canonical_slug

Rows where keep_or_redirect == REDIRECT become:
- Redirect from: /blogs/news/[redirect_from_slug]
- Redirect to: /blogs/news/[canonical_slug]

Use --canonical-slugs to process one topic cluster at a time.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List


DEFAULT_BLOG_PATH = "/blogs/news"
DEFAULT_OUTPUT_DIR = Path("ops/content/style-journal")
REQUIRED_COLUMNS = {"keep_or_redirect", "redirect_from_slug", "canonical_slug"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="Prompt 6 spreadsheet path (.csv, .tsv, or .xlsx)")
    parser.add_argument(
        "--canonical-slugs",
        default="",
        help="Comma-separated canonical slugs to limit execution to one or more topic clusters",
    )
    parser.add_argument(
        "--blog-path",
        default=DEFAULT_BLOG_PATH,
        help=f"Blog path prefix for redirects (default: {DEFAULT_BLOG_PATH})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Directory for generated outputs (default: {DEFAULT_OUTPUT_DIR})",
    )
    parser.add_argument(
        "--basename",
        default="blog-consolidation-redirects",
        help="Base filename for generated output files",
    )
    return parser.parse_args()


def clean(value: object) -> str:
    return str(value or "").strip()


def normalize_slug(value: str) -> str:
    return clean(value).strip("/")


def normalize_blog_path(value: str) -> str:
    path = "/" + clean(value).strip("/")
    return path.rstrip("/")


def load_csv_rows(path: Path, delimiter: str) -> List[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        return [dict(row) for row in reader]


def load_xlsx_rows(path: Path) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading .xlsx files requires openpyxl. Convert the sheet to CSV/TSV or install openpyxl."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [clean(value) for value in rows[0]]
    records: List[Dict[str, str]] = []
    for row in rows[1:]:
        record = {headers[index]: clean(row[index]) for index in range(min(len(headers), len(row)))}
        records.append(record)
    return records


def load_rows(path: Path) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv_rows(path, ",")
    if suffix == ".tsv":
        return load_csv_rows(path, "\t")
    if suffix == ".xlsx":
        return load_xlsx_rows(path)
    raise RuntimeError(f"Unsupported input format: {path.suffix}")


def normalize_record(record: Dict[str, str]) -> Dict[str, str]:
    normalized = {clean(key).strip().lower(): clean(value) for key, value in record.items()}
    return normalized


def ensure_required_columns(rows: Iterable[Dict[str, str]]) -> None:
    try:
        first_row = next(iter(rows))
    except StopIteration as exc:
        raise RuntimeError("The consolidation sheet is empty.") from exc

    missing = REQUIRED_COLUMNS - set(first_row.keys())
    if missing:
        raise RuntimeError(f"Missing required columns: {', '.join(sorted(missing))}")


def write_csv(path: Path, rows: List[Dict[str, str]], fieldnames: List[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_jsonl(path: Path, rows: List[Dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=True) + "\n")


def main() -> int:
    args = parse_args()
    raw_rows = load_rows(args.input.expanduser())
    normalized_rows = [normalize_record(row) for row in raw_rows]
    ensure_required_columns(normalized_rows)

    requested_canonicals = {
        normalize_slug(part) for part in clean(args.canonical_slugs).split(",") if normalize_slug(part)
    }
    blog_path = normalize_blog_path(args.blog_path)

    redirect_rows: List[Dict[str, str]] = []
    api_rows: List[Dict[str, Dict[str, str]]] = []
    detail_rows: List[Dict[str, str]] = []
    cluster_counts: Counter[str] = Counter()
    source_targets: Dict[str, str] = {}

    for row in normalized_rows:
        decision = clean(row.get("keep_or_redirect", "")).upper()
        if decision != "REDIRECT":
            continue

        redirect_from_slug = normalize_slug(row.get("redirect_from_slug", ""))
        canonical_slug = normalize_slug(row.get("canonical_slug", ""))
        if requested_canonicals and canonical_slug not in requested_canonicals:
            continue
        if not redirect_from_slug or not canonical_slug:
            continue

        source = f"{blog_path}/{redirect_from_slug}"
        target = f"{blog_path}/{canonical_slug}"
        if source == target:
            continue

        previous_target = source_targets.get(source)
        if previous_target and previous_target != target:
            raise RuntimeError(f"Conflicting targets for {source}: {previous_target} vs {target}")
        source_targets[source] = target

        cluster_counts[canonical_slug] += 1
        redirect_rows.append({"Redirect from": source, "Redirect to": target})
        api_rows.append({"redirect": {"path": source, "target": target}})
        detail_rows.append(
            {
                "keep_or_redirect": decision,
                "redirect_from_slug": redirect_from_slug,
                "canonical_slug": canonical_slug,
                "redirect_from_path": source,
                "redirect_to_path": target,
            }
        )

    if not redirect_rows:
        print("No redirect rows matched the requested scope.", file=sys.stderr)
        return 1

    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / f"{args.basename}.csv"
    jsonl_path = output_dir / f"{args.basename}.jsonl"
    details_path = output_dir / f"{args.basename}-details.csv"

    write_csv(csv_path, redirect_rows, ["Redirect from", "Redirect to"])
    write_jsonl(jsonl_path, api_rows)
    write_csv(
        details_path,
        detail_rows,
        [
            "keep_or_redirect",
            "redirect_from_slug",
            "canonical_slug",
            "redirect_from_path",
            "redirect_to_path",
        ],
    )

    print(f"input={args.input}")
    print(f"redirect_rows={len(redirect_rows)}")
    print(f"clusters={len(cluster_counts)}")
    print(f"csv={csv_path}")
    print(f"jsonl={jsonl_path}")
    print(f"details={details_path}")
    for canonical_slug, count in sorted(cluster_counts.items()):
        print(f"cluster[{canonical_slug}]={count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
